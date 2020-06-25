
import openpyxl

from pathlib import Path

xlsx_file = Path('','SUBTLEX-US frequency.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)

sheet = wb_obj.active

col_names = []

last_column = sheet.max_column-1
for i, row in enumerate(sheet.iter_rows(values_only=True)):
	if i==0:
		continue
	print(row[0])
	print(round(float(row[last_column]),1))
