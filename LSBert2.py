#!/usr/bin/python
# -*- coding: UTF-8 -*-



import argparse
import csv
import os
import random
import math
import sys
import re

from pytorch_pretrained_bert.tokenization import BertTokenizer
from pytorch_pretrained_bert.modeling import BertModel, BertForMaskedLM

from sklearn.metrics.pairwise import cosine_similarity as cosine

from scipy.special import softmax

import openpyxl

from pathlib import Path

from PPDB import Ppdb
from nltk.tokenize import word_tokenize

# OPTIONAL: if you want to have more information on what's happening, activate the logger as follows
import numpy as np
import torch
import nltk

from nltk.stem import PorterStemmer

class InputFeatures(object):
    """A single set of features of data."""

    def __init__(self, unique_id, tokens, input_ids, input_mask, input_type_ids):
        self.unique_id = unique_id
        self.tokens = tokens
        self.input_ids = input_ids
        self.input_mask = input_mask
        self.input_type_ids = input_type_ids

def convert_sentence_to_token(sentence, seq_length, tokenizer):

    tokenized_text = tokenizer.tokenize(sentence.lower())

    assert len(tokenized_text) < seq_length-2

    nltk_sent = nltk.word_tokenize(sentence.lower())

    position2 = []

    token_index = 0

    start_pos =  len(tokenized_text)  + 2

    pre_word = ""

    for i,word in enumerate(nltk_sent):

        if word=="n't" and pre_word[-1]=="n":
            word = "'t"

        if tokenized_text[token_index]=="\"":
            len_token = 2
        else:
            len_token = len(tokenized_text[token_index])

        if tokenized_text[token_index]==word or len_token>=len(word):
            position2.append(start_pos+token_index)
            pre_word = tokenized_text[token_index]

            token_index += 1
        else:
            new_pos = []
            new_pos.append(start_pos+token_index)

            new_word = tokenized_text[token_index]

            while new_word != word:

                token_index += 1

                new_word += tokenized_text[token_index].replace('##','')

                new_pos.append(start_pos+token_index)

                if len(new_word)==len(word):
                    break
            token_index += 1
            pre_word = new_word
           
            position2.append(new_pos)
       
    return tokenized_text, nltk_sent, position2

def convert_whole_word_to_feature(tokens_a, mask_position, seq_length, tokenizer, prob_mask):
    """Loads a data file into a list of `InputFeature`s."""

    #tokens_a = tokenizer.tokenize(sentence)
    #print(mask_position)
    #print("Convert_whole_word_to_feature")
    #print(tokens_a)

    tokens = []
    input_type_ids = []
    tokens.append("[CLS]")
    input_type_ids.append(0)

    len_tokens = len(tokens_a)
    first_sentence_mask_random = random.sample(range(0,len_tokens), int(prob_mask*len_tokens))

    mask_index = []

    for mask_pos in mask_position:
        mask_index.append(mask_pos-len_tokens-2)

    for i in range(len_tokens):

        if i in mask_index:
            tokens.append(tokens_a[i])
        elif i in first_sentence_mask_random:
            tokens.append('[MASK]')
        else:
            tokens.append(tokens_a[i])
        input_type_ids.append(0)
    
    tokens.append("[SEP]")
    input_type_ids.append(0)

    for token in tokens_a:
        tokens.append(token)
        input_type_ids.append(1)

    tokens.append("[SEP]")
    input_type_ids.append(1)

    true_word = ''
    index = 0
    count = 0
    mask_position_length = len(mask_position)

    while count in range(mask_position_length):
        index = mask_position_length - 1 - count

        pos = mask_position[index]
        if index == 0:
            tokens[pos] = '[MASK]'
        else:
            del tokens[pos]
            del input_type_ids[pos]

        count += 1

    #print(tokens)

    input_ids = tokenizer.convert_tokens_to_ids(tokens)

        # The mask has 1 for real tokens and 0 for padding tokens. Only real
        # tokens are attended to.
    input_mask = [1] * len(input_ids)

        # Zero-pad up to the sequence length.
    while len(input_ids) < seq_length:
        input_ids.append(0)
        input_mask.append(0)
        input_type_ids.append(0)

    assert len(input_ids) == seq_length
    assert len(input_mask) == seq_length
    assert len(input_type_ids) == seq_length

      
    return InputFeatures(unique_id=0,  tokens=tokens, input_ids=input_ids,input_mask=input_mask,input_type_ids=input_type_ids)
    

def convert_token_to_feature(tokens_a, mask_position, seq_length, tokenizer, prob_mask):
    """Loads a data file into a list of `InputFeature`s."""

    #tokens_a = tokenizer.tokenize(sentence)
    #print(mask_position)
    #print("----------")
    #print(tokens_a)

    tokens = []
    input_type_ids = []
    tokens.append("[CLS]")
    input_type_ids.append(0)

    len_tokens = len(tokens_a)
    #print("length of tokens: ", len_tokens)

    first_sentence_mask_random = random.sample(range(0,len_tokens), int(prob_mask*len_tokens))

    for i in range(len_tokens):

        if i==(mask_position-len_tokens-2):
            tokens.append(tokens_a[i])
        elif i in first_sentence_mask_random:
            tokens.append('[MASK]')
        else:
            tokens.append(tokens_a[i])
        input_type_ids.append(0)
    
    tokens.append("[SEP]")
    input_type_ids.append(0)

    for token in tokens_a:
        tokens.append(token)
        input_type_ids.append(1)

    tokens.append("[SEP]")
    input_type_ids.append(1)

    true_word = ''
    true_word = tokens[mask_position]
    tokens[mask_position] =  '[MASK]'

    #print(tokens)

    input_ids = tokenizer.convert_tokens_to_ids(tokens)

        # The mask has 1 for real tokens and 0 for padding tokens. Only real
        # tokens are attended to.
    input_mask = [1] * len(input_ids)

        # Zero-pad up to the sequence length.
    while len(input_ids) < seq_length:
        input_ids.append(0)
        input_mask.append(0)
        input_type_ids.append(0)

    if len(input_ids) > seq_length:
        print(len(input_ids))

    assert len(input_ids) == seq_length
    assert len(input_mask) == seq_length
    assert len(input_type_ids) == seq_length

      
    return InputFeatures(unique_id=0,  tokens=tokens, input_ids=input_ids,input_mask=input_mask,input_type_ids=input_type_ids)
    

def getWordmap(wordVecPath):
    words=[]
    We = []
    f = open(wordVecPath,'r')
    lines = f.readlines()

    for (n,line) in enumerate(lines):
        if (n == 0) :
            print(line)
            continue
        word, vect = line.rstrip().split(' ', 1)
                    
        vect = np.fromstring(vect, sep=' ')
                
        We.append(vect)

        words.append(word)

        #if(n==200000):
        #    break
    f.close()       
    return (words, We)


def getWordCount(word_count_path):
    word2count = {}
    xlsx_file = Path('',word_count_path)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    last_column = sheet.max_column-1
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i==0:
            continue
        word2count[row[0]] = round(float(row[last_column]),3)
        
    return word2count

def read_eval_index_dataset(data_path, is_label=True):
    sentences=[]
    mask_words = []
    mask_labels = []

    with open(data_path, "r", encoding='ISO-8859-1') as reader:
        while True:
            line = reader.readline()
            
            if not line:
                break
            
            sentence,words = line.strip().split('\t',1)
                #print(sentence)
            mask_word,labels = words.strip().split('\t',1)
            label = labels.split('\t')
                
            sentences.append(sentence)
            mask_words.append(mask_word)
                
            one_labels = []
            for la in label[1:]:
                if la not in one_labels:
                    la_id,la_word = la.split(':')
                    one_labels.append(la_word)
                
                #print(mask_word, " ---",one_labels)
            mask_labels.append(one_labels)
            
    return sentences,mask_words,mask_labels

def read_eval_dataset(data_path, is_label=True):
    sentences=[]
    mask_words = []
    mask_labels = []
    id = 0

    with open(data_path, "r", encoding='ISO-8859-1') as reader:
        while True:
            line = reader.readline()
            if is_label:
                id += 1
                if id==1:
                    continue
                if not line:
                    break
                sentence,words = line.strip().split('\t',1)
                #print(sentence)
                mask_word,labels = words.strip().split('\t',1)
                label = labels.split('\t')
                
                sentences.append(sentence)
                mask_words.append(mask_word)
                
                one_labels = []
                for la in label:
                    if la not in one_labels:
                        one_labels.append(la)
                
                #print(mask_word, " ---",one_labels)
                    
                mask_labels.append(one_labels)
            else:
                if not line:
                    break
                #print(line)
                sentence,mask_word = line.strip().split('\t')
                sentences.append(sentence)
                mask_words.append(mask_word)
    return sentences,mask_words,mask_labels

def BERT_candidate_generation(source_word, pre_tokens, pre_scores, ps, num_selection=10):

    cur_tokens=[]
   

    source_stem = ps.stem(source_word)

    assert num_selection<=len(pre_tokens)

    for i in range(len(pre_tokens)):
        token = pre_tokens[i]
     
        if token[0:2]=="##":
            continue

        if(token==source_word):
            continue

        token_stem = ps.stem(token)

        if(token_stem == source_stem):
            continue

        if (len(token_stem)>=3) and (token_stem[:3]==source_stem[:3]):
            continue

        cur_tokens.append(token)
        

        if(len(cur_tokens)==num_selection):
            break
    
    if(len(cur_tokens)==0):
        cur_tokens = pre_tokens[0:num_selection+1]
        

    assert len(cur_tokens)>0       

    return cur_tokens

def cross_entropy_word(X,i,pos):
    
    #print(X)
    #print(X[0,2,3])
    X = softmax(X,axis=1)
    loss = 0
    loss -= np.log10(X[i,pos])
    return loss


def get_score(sentence,tokenizer,maskedLM):
    tokenize_input = tokenizer.tokenize(sentence)

    len_sen = len(tokenize_input)

    START_TOKEN = '[CLS]'
    SEPARATOR_TOKEN = '[SEP]'

    tokenize_input.insert(0, START_TOKEN)
    tokenize_input.append(SEPARATOR_TOKEN)

    input_ids = tokenizer.convert_tokens_to_ids(tokenize_input)

    #tensor_input = torch.tensor([tokenizer.convert_tokens_to_ids(tokenize_input)])
    #print("tensor_input")
    #print(tensor_input)
    #tensor_input = tensor_input.to('cuda')
    sentence_loss = 0
    
    for i,word in enumerate(tokenize_input):

        if(word == START_TOKEN or word==SEPARATOR_TOKEN):
            continue

        orignial_word = tokenize_input[i]
        tokenize_input[i] = '[MASK]'
        #print(tokenize_input)
        mask_input = torch.tensor([tokenizer.convert_tokens_to_ids(tokenize_input)])
        #print(mask_input)
        mask_input = mask_input.to('cuda')
        with torch.no_grad():
            att, pre_word =maskedLM(mask_input)
        word_loss = cross_entropy_word(pre_word[0].cpu().numpy(),i,input_ids[i])
        sentence_loss += word_loss
        #print(word_loss)
        tokenize_input[i] = orignial_word
        
    return np.exp(sentence_loss/len_sen)


def LM_score(source_word,source_context,substitution_selection,tokenizer,maskedLM):
    #source_index = source_context.index(source_word)

    source_sentence = ''

    for context in source_context:
        source_sentence += context + " "
    
    source_sentence = source_sentence.strip()
    #print(source_sentence)
    LM = []

    source_loss = get_score(source_sentence,tokenizer,maskedLM)

    for substibution in substitution_selection:
        
        sub_sentence = source_sentence.replace(source_word,substibution)

        
        #print(sub_sentence)
        score = get_score(sub_sentence,tokenizer,maskedLM)

        #print(score)
        LM.append(score)

    return LM,source_loss


def preprocess_SR(source_word, substitution_selection, fasttext_dico, fasttext_emb, word_count):
    ss = []
    ##ss_score=[]
    sis_scores=[]
    count_scores=[]

    isFast = True

    if(source_word not in fasttext_dico):
        isFast = False
    else:
        source_emb = fasttext_emb[fasttext_dico.index(source_word)].reshape(1,-1)

    #ss.append(source_word)

    for sub in substitution_selection:

        if sub not in word_count:
            continue
        else:
            sub_count = word_count[sub]

        if(sub_count<=3):
            continue

        #if sub_count<source_count:
         #   continue
        if isFast:
            if sub not in fasttext_dico:
                continue

            token_index_fast = fasttext_dico.index(sub)
            sis = cosine(source_emb, fasttext_emb[token_index_fast].reshape(1,-1))

            #if sis<0.35:
            #    continue
            sis_scores.append(sis)

        ss.append(sub)
        count_scores.append(sub_count)

    return ss,sis_scores,count_scores

def compute_context_sis_score(source_word, sis_context, substitution_selection, fasttext_dico, fasttext_emb):
    context_sis = []

    word_context = []

    

    for con in sis_context:
        if con==source_word or (con not in fasttext_dico):
            continue

        word_context.append(con)

    if len(word_context)!=0:
        for sub in substitution_selection:
            sub_emb = fasttext_emb[fasttext_dico.index(sub)].reshape(1,-1)
            all_sis = 0
            for con in word_context:
                token_index_fast = fasttext_dico.index(con)
                all_sis += cosine(sub_emb, fasttext_emb[token_index_fast].reshape(1,-1))

            context_sis.append(all_sis/len(word_context))
    else:
        for i in range(len(substitution_selection)):
            context_sis.append(len(substitution_selection)-i)

            
    return context_sis


def substitution_ranking(source_word, source_context, substitution_selection, fasttext_dico, fasttext_emb, word_count, ssPPDB, tokenizer, maskedLM, lables):

    ss,sis_scores,count_scores=preprocess_SR(source_word, substitution_selection, fasttext_dico, fasttext_emb, word_count)

    #print(ss)
    if len(ss)==0:
        return source_word

    if len(sis_scores)>0:
        seq = sorted(sis_scores,reverse = True )
        sis_rank = [seq.index(v)+1 for v in sis_scores]
    
    rank_count = sorted(count_scores,reverse = True )
    
    count_rank = [rank_count.index(v)+1 for v in count_scores]
  
    lm_score,source_lm = LM_score(source_word,source_context,ss,tokenizer,maskedLM)

    rank_lm = sorted(lm_score)
    lm_rank = [rank_lm.index(v)+1 for v in lm_score]
    

    bert_rank = []
    ppdb_rank =[]
    for i in range(len(ss)):
        bert_rank.append(i+1)

        if ss[i] in ssPPDB:
        	ppdb_rank.append(1)
        else:
        	ppdb_rank.append(len(ss)/3)

    if len(sis_scores)>0:
        all_ranks = [bert+sis+count+LM+ppdb  for bert,sis,count,LM,ppdb in zip(bert_rank,sis_rank,count_rank,lm_rank,ppdb_rank)]
    else:
        all_ranks = [bert+count+LM+ppdb  for bert,count,LM,ppdb in zip(bert_rank,count_rank,lm_rank,ppdb_rank)]
    #all_ranks = [con for con in zip(context_rank)]


    pre_index = all_ranks.index(min(all_ranks))

    #return ss[pre_index]

    pre_count = count_scores[pre_index]

    if source_word in word_count:
    	source_count = word_count[source_word]
    else:
    	source_count = 0

    pre_lm = lm_score[pre_index]

    #print(lm_score)
    #print(source_lm)
    #print(pre_lm)


    #pre_word = ss[pre_index]


    if source_lm>pre_lm or pre_count>source_count:
    	pre_word = ss[pre_index]
    else:
    	pre_word = source_word

    return pre_word


def evaulation_SS_scores(ss,labels):
    assert len(ss)==len(labels)

    potential = 0
    instances = len(ss)
    precision = 0
    precision_all = 0
    recall = 0
    recall_all = 0

    for i in range(len(ss)):

        one_prec = 0
        
        common = list(set(ss[i]).intersection(labels[i]))

        if len(common)>=1:
            potential +=1
        precision += len(common)
        recall += len(common)
        precision_all += len(ss[i])
        recall_all += len(labels[i])

    potential /=  instances
    precision /= precision_all
    recall /= recall_all
    F_score = 2*precision*recall/(precision+recall)

    return potential,precision,recall,F_score


def evaulation_pipeline_scores(substitution_words,source_words,gold_words):

    instances = len(substitution_words)
    precision = 0
    accuracy = 0
    changed_proportion = 0

    for sub, source, gold in zip(substitution_words,source_words,gold_words):
        if sub==source or (sub in gold):
            precision += 1
        if sub!=source and (sub in gold):
            accuracy += 1
        if sub!=source:
            changed_proportion += 1

    return precision/instances,accuracy/instances,changed_proportion/instances




def extract_context(words, mask_index, window):
    #extract 7 words around the content word

    length = len(words)

    half = int(window/2)

    assert mask_index>=0 and mask_index<length

    context = ""

    if length<=window:
        context = words
    elif mask_index<length-half and mask_index>=half:
        context = words[mask_index-half:mask_index+half+1]
    elif mask_index<half:
        context = words[0:window]
    elif mask_index>=length-half:
        context = words[length-window:length]
    else:
        print("Wrong!")

    return context

def preprocess_tag(tag):
    if tag[0] =="V" or tag[0]=="N":
        return tag
    if tag[0]=="R":
        return "r"
    if tag[0]=="J" or tag[0]=="I":
        return 'a'
    else:
        return 's'  

def main():
    parser = argparse.ArgumentParser()

    ## Required parameters
    
    parser.add_argument("--eval_dir",
                        default=None,
                        type=str,
                        required=True,
                        help="The evaluation data dir.")
    parser.add_argument("--bert_model", default=None, type=str, required=True,
                        help="Bert pre-trained model selected in the list: bert-base-uncased, "
                        "bert-large-uncased, bert-base-cased, bert-large-cased, bert-base-multilingual-uncased, "
                        "bert-base-multilingual-cased, bert-base-chinese.")

    parser.add_argument("--output_SR_file",
                        default=None,
                        type=str,
                        required=True,
                        help="The output directory of writing substitution selection.")
    parser.add_argument("--word_embeddings",
                        default=None,
                        type=str,
                        required=True,
                        help="The path of word embeddings")
    parser.add_argument("--word_frequency",
                        default=None,
                        type=str,
                        required=True,
                        help="The path of word frequency.")
    ## Other parameters
    parser.add_argument("--cache_dir",
                        default="",
                        type=str,
                        help="Where do you want to store the pre-trained models downloaded from s3")

    parser.add_argument("--max_seq_length",
                        default=250,
                        type=int,
                        help="The maximum total input sequence length after WordPiece tokenization. \n"
                             "Sequences longer than this will be truncated, and sequences shorter \n"
                             "than this will be padded.")

    parser.add_argument("--do_eval",
                        action='store_true',
                        help="Whether to run eval on the dev set.")
    parser.add_argument("--do_lower_case",
                        action='store_true',
                        help="Set this flag if you are using an uncased model.")

    parser.add_argument("--eval_batch_size",
                        default=8,
                        type=int,
                        help="Total batch size for eval.")
    parser.add_argument("--num_selections",
                        default=20,
                        type=int,
                        help="Total number of training epochs to perform.")
    parser.add_argument("--num_eval_epochs",
                        default=1,
                        type=int,
                        help="Total number of training epochs to perform.")

    parser.add_argument("--prob_mask",
                        default=0.5,
                        type=float,
                        help="Proportion of the masked words in first sentence. "
                             "E.g., 0.1 = 10%% of training.")
    parser.add_argument("--ppdb",
                        default="./ppdb-2.0-tldr",
                        type=str,
                        required=True,
                        help="The path of word frequency.")
    parser.add_argument("--no_cuda",
                        action='store_true',
                        help="Whether not to use CUDA when available")
    parser.add_argument("--local_rank",
                        type=int,
                        default=-1,
                        help="local_rank for distributed training on gpus")
    parser.add_argument('--seed',
                        type=int,
                        default=42,
                        help="random seed for initialization")
    args = parser.parse_args()



    if args.local_rank == -1 or args.no_cuda:
        device = torch.device("cuda" if torch.cuda.is_available() and not args.no_cuda else "cpu")
        n_gpu = torch.cuda.device_count()
    else:
        torch.cuda.set_device(args.local_rank)
        device = torch.device("cuda", args.local_rank)
        n_gpu = 1
        # Initializes the distributed backend which will take care of sychronizing nodes/GPUs
        torch.distributed.init_process_group(backend='nccl')

    random.seed(args.seed)
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    if n_gpu > 0:
        torch.cuda.manual_seed_all(args.seed)

    if  not args.do_eval:
        raise ValueError("At least `do_eval` must be True.")

    tokenizer = BertTokenizer.from_pretrained(args.bert_model, do_lower_case=args.do_lower_case)


    # Prepare model
    model = BertForMaskedLM.from_pretrained(args.bert_model,output_attentions=True)
    
    model.to(device)

    output_sr_file = open(args.output_SR_file,"a+")

    print("Loading embeddings ...")
    
    wordVecPath = args.word_embeddings
    #wordVecPath = "/media/qiang/ee63f41d-4004-44fe-bcfd-522df9f2eee8/glove.840B.300d.txt"

    fasttext_dico, fasttext_emb = getWordmap(wordVecPath)

    #stopword = set(stopwords.words('english'))
    word_count_path = args.word_frequency
    #word_count_path = "word_frequency_wiki.txt"
    word_count = getWordCount(word_count_path)

    ps = PorterStemmer()

    print("loading PPDB ...")
    ppdb_path = args.ppdb
    ppdb_model = Ppdb(ppdb_path)

    CGBERT = []

    substitution_words = []
   
    num_selection = args.num_selections

    window_context = 11

    if args.do_eval and (args.local_rank == -1 or torch.distributed.get_rank() == 0):
        
     
        fileName = args.eval_dir.split('/')[-1][:-4]
        if fileName=='lex.mturk':
            eval_examples, mask_words, mask_labels = read_eval_dataset(args.eval_dir)
        else:
            eval_examples, mask_words, mask_labels = read_eval_index_dataset(args.eval_dir)

       
        eval_size = len(eval_examples)
        print("***** Running evaluation *****")
        print("  Num examples = %d", eval_size)
        #logger.info("  Batch size = %d", args.eval_batch_size)

        model.eval()

        for i in range(eval_size):

            print('Sentence {} rankings: '.format(i))
            #output_sr_file.write(str(i))
            #output_sr_file.write(' sentence: ')
            #output_sr_file.write('\n')
            print(eval_examples[i])
            print(mask_words[i])
            
            tokens, words, position = convert_sentence_to_token(eval_examples[i], args.max_seq_length, tokenizer)

            assert len(words)==len(position)

            mask_index = words.index(mask_words[i])

            mask_context = extract_context(words,mask_index,window_context)

            len_tokens = len(tokens)

            mask_position = position[mask_index]
 
            if isinstance(mask_position,list):
                feature = convert_whole_word_to_feature(tokens, mask_position, args.max_seq_length, tokenizer, args.prob_mask)
            else:
                feature = convert_token_to_feature(tokens, mask_position, args.max_seq_length, tokenizer, args.prob_mask)

            tokens_tensor = torch.tensor([feature.input_ids])

            token_type_ids = torch.tensor([feature.input_type_ids])

            attention_mask = torch.tensor([feature.input_mask])

            tokens_tensor = tokens_tensor.to('cuda')
            token_type_ids = token_type_ids.to('cuda')
            attention_mask = attention_mask.to('cuda')

                # Predict all tokens
            with torch.no_grad():
                all_attentions,prediction_scores = model(tokens_tensor, token_type_ids,attention_mask)

            if isinstance(mask_position,list):
                predicted_top = prediction_scores[0, mask_position[0]].topk(80)
            else:
                predicted_top = prediction_scores[0, mask_position].topk(80)
                #print(predicted_top[0].cpu().numpy())
            pre_tokens = tokenizer.convert_ids_to_tokens(predicted_top[1].cpu().numpy())
            
            #print(predicted_top[0].cpu().numpy())

            sentence = eval_examples[i].lower()
            words = word_tokenize(sentence)

            words_tag = nltk.pos_tag(words)

            complex_word_index = words.index(mask_words[i])

            complex_word_tag = words_tag[complex_word_index][1]

   

            complex_word_tag = preprocess_tag(complex_word_tag)
            
            cgPPDB = ppdb_model.predict(mask_words[i],complex_word_tag)

            cgBERT = BERT_candidate_generation(mask_words[i], pre_tokens, predicted_top[0].cpu().numpy(), ps, args.num_selections)

            print(cgBERT)
            
            CGBERT.append(cgBERT)
          
            pre_word = substitution_ranking(mask_words[i], mask_context, cgBERT, fasttext_dico, fasttext_emb,word_count,cgPPDB,tokenizer,model,mask_labels[i])


            substitution_words.append(pre_word)

        
        potential,precision,recall,F_score=evaulation_SS_scores(CGBERT, mask_labels)
        print("The score of evaluation for BERT candidate generation")
        print(potential,precision,recall,F_score)

        output_sr_file.write(str(args.num_selections))
        output_sr_file.write('\t')
        output_sr_file.write(str(potential))
        output_sr_file.write('\t')
        output_sr_file.write(str(precision))
        output_sr_file.write('\t')
        output_sr_file.write(str(recall))
        output_sr_file.write('\t')
        output_sr_file.write(str(F_score))
        output_sr_file.write('\t')
        

        precision,accuracy,changed_proportion=evaulation_pipeline_scores(substitution_words, mask_words, mask_labels)
       	print("The score of evaluation for full LS pipeline")
        print(precision,accuracy,changed_proportion)
        output_sr_file.write(str(precision))
        output_sr_file.write('\t')
        output_sr_file.write(str(accuracy))
        output_sr_file.write('\t')
        output_sr_file.write(str(changed_proportion))
        output_sr_file.write('\n')

        output_sr_file.close()

if __name__ == "__main__":
    main()

